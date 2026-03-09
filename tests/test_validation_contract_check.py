"""
Unit tests for pipelines.validation.contract_check: CLI exit codes and preflight_validation_inputs.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from legacy.legacy_pipelines.contracts.validation_policy_contract import (
    DEFAULT_HIGHLIGHTED,
    ExpectedColumnsConfig,
    FailFastConfig,
    NormalizationConfig,
    ToleranceConfig,
    ValidationPolicy,
    WorkbookConfig,
)
from legacy.legacy_pipelines.validation.contract_check import (
    preflight_validation_inputs,
    print_validation_summary,
)


def _make_policy(workbook_path: str = "data/expected_summary.xlsx") -> ValidationPolicy:
    return ValidationPolicy(
        workbook=WorkbookConfig(
            path=workbook_path,
            sheet="DATA SUMMARY",
            month_column="Month",
            month_format=None,
        ),
        expected_columns=ExpectedColumnsConfig(
            asset_growth_rate="AssetGrowth",
            organic_growth_rate="OrganicGrowth",
            external_market_growth_rate="ExternalMarketGrowth",
        ),
        normalization=NormalizationConfig(
            percent_to_decimal=False,
            percent_scale=100.0,
            month_align="month_end",
            timezone_naive=True,
        ),
        tolerance=ToleranceConfig(abs_tol=0.0, rel_tol=0.0),
        fail_fast=FailFastConfig(
            max_mismatched_months=10,
            max_deviation=1.0,
            fail_on_missing_months=True,
        ),
        highlighted=DEFAULT_HIGHLIGHTED,
    )


def test_print_validation_summary(capsys: pytest.CaptureFixture[str]) -> None:
    policy = _make_policy()
    print_validation_summary(policy)
    out = capsys.readouterr().out
    assert "workbook.path" in out
    assert "DATA SUMMARY" in out
    assert "Month" in out
    assert "AssetGrowth" in out
    assert "asset_growth_rate" in out
    assert "percent_to_decimal" in out
    assert "max_mismatched_months" in out


def test_preflight_workbook_missing() -> None:
    policy = _make_policy(workbook_path="nonexistent/workbook.xlsx")
    with pytest.raises(FileNotFoundError, match="Validation workbook not found"):
        preflight_validation_inputs(policy, root=PROJECT_ROOT, verbose=False)


def test_preflight_parquet_missing(tmp_path: Path) -> None:
    # Workbook exists (create minimal Excel); parquet does not.
    from openpyxl import Workbook
    xlsx = tmp_path / "summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA SUMMARY"
    ws.append(["Month", "AssetGrowth", "OrganicGrowth", "ExternalMarketGrowth"])
    ws.append(["2024-01-01", 1.0, 0.5, 1.0])
    wb.save(xlsx)
    policy = _make_policy(workbook_path=str(xlsx))
    with pytest.raises(FileNotFoundError, match="Metrics artifact not found"):
        preflight_validation_inputs(policy, root=tmp_path, verbose=False)


def test_preflight_sheet_not_found(tmp_path: Path) -> None:
    from openpyxl import Workbook
    xlsx = tmp_path / "summary.xlsx"
    wb = Workbook()
    wb.active.title = "OtherSheet"
    wb.save(xlsx)
    policy = _make_policy(workbook_path=str(xlsx))
    with pytest.raises(ValueError, match="Sheet .* not found") as exc_info:
        preflight_validation_inputs(policy, root=tmp_path, verbose=False)
    assert "Available sheets" in str(exc_info.value)


def test_preflight_columns_missing(tmp_path: Path) -> None:
    from openpyxl import Workbook
    xlsx = tmp_path / "summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA SUMMARY"
    ws.append(["Month", "AssetGrowth"])  # missing OrganicGrowth, ExternalMarketGrowth
    wb.save(xlsx)
    policy = _make_policy(workbook_path=str(xlsx))
    with pytest.raises(ValueError, match="Required columns missing") as exc_info:
        preflight_validation_inputs(policy, root=tmp_path, verbose=False)
    assert "Required:" in str(exc_info.value)
    assert "Found:" in str(exc_info.value)
    assert "Missing:" in str(exc_info.value)


def test_preflight_success(tmp_path: Path) -> None:
    from openpyxl import Workbook
    xlsx = tmp_path / "summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA SUMMARY"
    ws.append(["Month", "AssetGrowth", "OrganicGrowth", "ExternalMarketGrowth"])
    ws.append(["2024-01-01", 1.0, 0.5, 1.0])
    wb.save(xlsx)
    parquet_dir = tmp_path / "curated"
    parquet_dir.mkdir(parents=True)
    (parquet_dir / "metrics_monthly.parquet").write_bytes(b"\x00")  # minimal placeholder
    policy = _make_policy(workbook_path=str(xlsx))
    preflight_validation_inputs(policy, root=tmp_path, verbose=False)  # no raise


def test_cli_missing_config_exits_nonzero() -> None:
    result = __import__("subprocess").run(
        [sys.executable, "-m", "pipelines.validation.contract_check", "--config", "nonexistent.yml"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
    )
    assert result.returncode != 0
    assert "Validation policy" in result.stderr or "not found" in result.stderr
